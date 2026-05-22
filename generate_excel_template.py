#!/usr/bin/env python3
"""
generate_excel_template.py
==========================
Genera una plantilla Excel (.xlsx) para alimentar el migrador del sitio
QR Groupe SEB desde Excel (en lugar de Google Sites).

La plantilla viene pre-llenada con los datos actuales del sitio publicado
(leídos de docs/migration_metadata.json), de modo que el usuario solo edita
lo que necesita y agrega nuevos TAGs.

Estructura del Excel:
  Hoja 1 "Datos":
    - TAG (obligatorio)               ej. 800-P-01
    - Categoría (obligatorio, dropdown) EQUIPOS | INSTRUMENTOS | TANQUES
    - Título (opcional, default = TAG)
    - Tipo de Equipo
    - Función
    - Marca / Modelo
    - Especificaciones
    - Ficha Técnica   (hyperlink nativo de Excel)
    - Curva           (hyperlink nativo de Excel)
  Hoja 2 "Instrucciones": cómo llenar el archivo.

Uso:
    python3 generate_excel_template.py
        -> genera plantilla_sitio.xlsx en la raíz del proyecto
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

REPO_ROOT = Path(__file__).parent
METADATA = REPO_ROOT / "docs" / "migration_metadata.json"
OUTPUT = REPO_ROOT / "plantilla_sitio.xlsx"

CATEGORIES = ["EQUIPOS", "INSTRUMENTOS", "TANQUES"]
SECTION_COLS = ["Tipo de Equipo", "Función", "Marca / Modelo", "Especificaciones"]
LINK_COLS = ["Ficha Técnica", "Curva"]
HEADERS = ["TAG", "Categoría", "Título"] + SECTION_COLS + LINK_COLS


def load_existing_data():
    """Lee migration_metadata.json y normaliza a filas de Excel."""
    if not METADATA.exists():
        print(f"⚠️  No se encontró {METADATA}. La plantilla saldrá vacía.")
        return []

    data = json.loads(METADATA.read_text(encoding="utf-8"))
    rows = []
    for p in data.get("pages", []):
        if p.get("is_index"):
            continue
        row = {
            "TAG": p["item_id"],
            "Categoría": p["category"],
            "Título": p["title"] if p["title"] != p["item_id"] else "",
        }
        # Mapear secciones a columnas
        section_map = {s["title"]: s.get("content", "") for s in p.get("sections", [])}
        for col in SECTION_COLS:
            row[col] = section_map.get(col, "")
        # Mapear links externos
        link_map = {l["text"]: l["url"] for l in p.get("links", [])}
        for col in LINK_COLS:
            row[col] = link_map.get(col, "")
        rows.append(row)
    return rows


def style_header(ws, n_cols):
    """Da formato a la fila de headers."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0969DA")
    border = Border(bottom=Side(style="medium", color="0550AE"))
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 28


def autosize_columns(ws, headers, rows):
    """Ajusta el ancho de cada columna al contenido."""
    for i, h in enumerate(headers, start=1):
        max_len = len(h)
        for r in rows:
            v = str(r.get(h, "") or "")
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 3, 12), 60)


def add_category_validation(ws, n_rows):
    """Agrega dropdown de categoría en la columna B desde fila 2."""
    formula = f'"{",".join(CATEGORIES)}"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False,
        showDropDown=False,
        errorTitle="Categoría inválida",
        error=f"Debe ser una de: {', '.join(CATEGORIES)}",
        promptTitle="Categoría",
        prompt=f"Selecciona: {', '.join(CATEGORIES)}",
    )
    ws.add_data_validation(dv)
    last_row = max(n_rows + 100, 500)  # buffer para agregar nuevas filas
    dv.add(f"B2:B{last_row}")


def write_data_sheet(wb, rows):
    """Escribe la hoja 'Datos'."""
    ws = wb.active
    ws.title = "Datos"

    # Headers
    for col_idx, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header(ws, len(HEADERS))

    # Filas
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(HEADERS, start=1):
            value = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Hyperlinks en columnas de links
            if h in LINK_COLS and value and value.startswith("http"):
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=value, display=h)
                cell.value = h  # el texto visible es "Ficha Técnica" / "Curva"
                cell.font = Font(color="0969DA", underline="single")

    # Freeze panes en fila 1 + columna A
    ws.freeze_panes = "C2"

    # Validación de categoría
    add_category_validation(ws, len(rows))

    # Ancho automático
    autosize_columns(ws, HEADERS, rows)


def write_instructions_sheet(wb):
    """Crea hoja con instrucciones de uso."""
    ws = wb.create_sheet("Instrucciones")

    title_font = Font(bold=True, size=14, color="0969DA")
    h2_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    code_font = Font(size=10, name="Consolas", color="333333")

    lines = [
        ("Plantilla de datos para el sitio QR Groupe SEB", title_font),
        ("", normal_font),
        ("Cada fila de la hoja 'Datos' es un TAG (un equipo/instrumento/tanque).", normal_font),
        ("Cada columna a partir de 'Tipo de Equipo' es una propiedad que aparecerá", normal_font),
        ("como sección colapsable en la ficha del TAG. Si dejas una celda vacía,", normal_font),
        ("esa sección NO se genera en la ficha.", normal_font),
        ("", normal_font),
        ("Columnas obligatorias", h2_font),
        ("• TAG          identificador único, ej. 100-P-01A", normal_font),
        ("• Categoría    EQUIPOS, INSTRUMENTOS o TANQUES (usa el dropdown)", normal_font),
        ("", normal_font),
        ("Columnas opcionales", h2_font),
        ("• Título           si lo dejas vacío, se usa el TAG como título", normal_font),
        ("• Tipo de Equipo   ej. Bomba de Coagulante", normal_font),
        ("• Función          ej. Dosificación de Coagulate en 200-MX-01", normal_font),
        ("• Marca / Modelo   texto libre, incluye serie si la tienes", normal_font),
        ("• Especificaciones texto libre (Tipo, Q, Potencia, Material, etc.)", normal_font),
        ("• Ficha Técnica    inserta hyperlink con Ctrl+K → pega la URL de Drive", normal_font),
        ("• Curva            ídem", normal_font),
        ("", normal_font),
        ("Cómo agregar columnas nuevas", h2_font),
        ("Si necesitas una nueva propiedad (ej. 'Capacidad', 'Material'), agrega", normal_font),
        ("una columna a la derecha de las existentes. El migrador la reconocerá", normal_font),
        ("automáticamente y la mostrará como una sección más en la ficha.", normal_font),
        ("", normal_font),
        ("Cómo agregar un TAG nuevo", h2_font),
        ("Solo agrega una fila al final de 'Datos' con su TAG y Categoría.", normal_font),
        ("El resto de columnas que dejes en blanco simplemente no se mostrarán.", normal_font),
        ("", normal_font),
        ("Regenerar el sitio desde este Excel", h2_font),
        ("Cuando termines de editar, guarda el archivo y ejecuta:", normal_font),
        ("    ./update_site_from_excel.sh", code_font),
        ("(script que se creará una vez confirmes la estructura)", normal_font),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = font
        c.alignment = Alignment(vertical="top", wrap_text=False)
    ws.column_dimensions["A"].width = 95
    ws.sheet_view.showGridLines = False


def main():
    rows = load_existing_data()
    print(f"📊 {len(rows)} TAGs cargados desde docs/migration_metadata.json")

    wb = Workbook()
    write_data_sheet(wb, rows)
    write_instructions_sheet(wb)

    wb.save(OUTPUT)
    print(f"✅ Plantilla generada: {OUTPUT}")
    print(f"   - {len(rows)} filas pre-llenadas")
    print(f"   - 2 hojas: Datos + Instrucciones")
    print(f"   - Dropdown de categoría en columna B")
    print(f"   - Hyperlinks preservados en columnas Ficha Técnica / Curva")


if __name__ == "__main__":
    main()
