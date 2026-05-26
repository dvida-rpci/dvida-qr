#!/usr/bin/env python3
"""
convert_fichas_to_template.py
=============================
Convierte un Excel de fichas técnicas (una hoja por TAG) en la plantilla
unificada compatible con el migrador del sitio (una sola hoja con todos
los TAGs en filas y propiedades en columnas).

Input:  NP00033 Fichas Técnicas Equipos 20250408 DG.xlsx (o similar)
Output: plantilla_sitio.xlsx (en la raíz del proyecto)

Reglas:
- Cada hoja del Excel origen es uno o varios TAGs.
- TAGs múltiples se expanden:
    "100-P-01 A_B"          -> "100-P-01A", "100-P-01B"
    "400-MX-01_02_03_04"    -> "400-MX-01", "400-MX-02", "400-MX-03", "400-MX-04"
- Categoría inferida del tipo en el TAG:
    TK, R, F, SD            -> TANQUES
    P, MX, M, C, PS         -> EQUIPOS
    AIT, FIT, SV, LT        -> INSTRUMENTOS
- Cada propiedad detectada en el archivo origen se vuelve una columna.
- Celdas vacías se mantienen vacías (no se genera esa sección en el sitio).
- Columnas 'Ficha Técnica' y 'Curva' se incluyen pero quedan vacías
  (este archivo no aporta links a documentos).
"""

import re
import shutil
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─────────────────────────────────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).parent
INPUT_DEFAULT = REPO_ROOT / "NP00033 Fichas Técnicas Equipos 20250408 DG.xlsx"
OUTPUT = REPO_ROOT / "plantilla_sitio.xlsx"
IMAGES_DIR = REPO_ROOT / "docs" / "assets" / "images"

# Imágenes que aparecen en MÁS de este umbral de hojas se consideran plantilla
# (logo, header, etc.) y se descartan. Solo se conservan las específicas del equipo.
# Umbral "plantilla común": una imagen que aparece en MÁS de N hojas se considera
# decorativa (logo/header del documento) y se descarta para todas las hojas.
# Subido de 5 → 30 (mayo 2026) porque imágenes de familias de equipos
# (p.ej. una bomba reutilizada en 14 fichas) caían como falsos positivos.
# Además, extract_images() tiene un fallback: si una hoja queda con 0 imágenes
# tras el filtro, conserva todas (mejor repetir que dejar una ficha sin imagen).
IMAGE_COMMON_THRESHOLD = 30

CATEGORIES = ["EQUIPOS", "INSTRUMENTOS", "TANQUES"]

# Mapping de prefijo de tipo (tras el primer número) → categoría
TYPE_TO_CATEGORY = {
    # TANQUES
    "TK": "TANQUES",
    "R": "TANQUES",
    "F": "TANQUES",
    "SD": "TANQUES",
    # EQUIPOS
    "P": "EQUIPOS",
    "MX": "EQUIPOS",
    "M": "EQUIPOS",
    "C": "EQUIPOS",
    "PS": "EQUIPOS",
    # INSTRUMENTOS
    "AIT": "INSTRUMENTOS",
    "FIT": "INSTRUMENTOS",
    "SV": "INSTRUMENTOS",
    "LT": "INSTRUMENTOS",
}

# Labels a IGNORAR (metadata administrativa, headers de sección)
# Nota: FLUIDO NO está aquí porque también es una propiedad legítima ("FLUIDO: ARnD").
# Se distingue por contexto: como block name está solo en col 0 sin valor a la derecha;
# como label tiene un valor en alguna celda posterior.
IGNORE_LABELS = {
    # Headers de bloque (estos NUNCA son propiedades)
    "GENERAL", "TANQUE", "EQUIPO", "INSTRUMENTO", "NOTAS", "FICHA TÉCNICA",
    # Metadata administrativa
    "PLANTA", "UBICACIÓN", "VERSIÓN", "FECHA", "ELABORÓ", "REVISÓ", "APROBÓ",
    "SOPORTE", "REV.", "REV",
    # Constantes redundantes que aparecen pegadas
    "STARND ETAPA 1 Y 2", "ITAGUI - ANTIOQUIA",
    "RIONEGRO, ANTIOQUIA", "STARND GROUPE SEB S.A.S",
    # Etiqueta del TAG (lo manejamos por separado vía sheet name)
    "TAG N°", "TAG N", "TAG",
    # Header decorativo de hojas auxiliares
    "BOMBA DE AGUA",
}


# ─────────────────────────────────────────────────────────────────────────
# Helpers de TAG
# ─────────────────────────────────────────────────────────────────────────
def expand_tags_from_sheet_name(sheet_name: str) -> list[str]:
    """Expande nombres de hoja a una lista de TAGs individuales."""
    name = sheet_name.strip()

    # Patrón "XXX A_B" o "XXX A/B" → XXXA, XXXB
    m = re.match(r"^(.+?)\s*A[_/]B\s*$", name)
    if m:
        base = m.group(1).strip()
        return [f"{base}A", f"{base}B"]

    # Patrón "XXX-NN_MM_OO_PP" → XXX-NN, XXX-MM, ...
    m = re.match(r"^(.+?-)(\d+(?:_\d+)+)$", name)
    if m:
        prefix, nums = m.group(1), m.group(2)
        return [f"{prefix}{n}" for n in nums.split("_")]

    # Sin patrón → TAG único
    return [name]


def infer_category(tag: str) -> str:
    """Deriva la categoría de un TAG por su tipo (XXX-TYPE-NN)."""
    parts = tag.split("-")
    if len(parts) < 2:
        return "EQUIPOS"  # fallback
    # parts[1] suele ser el tipo (TK, P, MX, M, C, PS, AIT, FIT, SV, LT, R, F, SD)
    type_code = parts[1].rstrip("0123456789 ").strip()
    return TYPE_TO_CATEGORY.get(type_code, "EQUIPOS")


# ─────────────────────────────────────────────────────────────────────────
# Extracción de pares (label, value)
# ─────────────────────────────────────────────────────────────────────────
def normalize_label(label: str) -> str:
    """Limpia espacios duplicados y caracteres raros del label."""
    s = re.sub(r"\s+", " ", label.strip())
    return s


def is_valid_label(label: str) -> bool:
    """Filtra labels que no son propiedades reales (metadata, headers, etc.)."""
    if not label or len(label) < 2 or len(label) > 80:
        return False
    upper = label.upper().strip()
    if upper in IGNORE_LABELS:
        return False
    # Texto que es una fecha
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", label):
        return False
    # Iniciales tipo "DG", "WO", "AB" (metadata de Elaboró/Revisó/Aprobó)
    # - 2-3 chars, todos mayúsculas, solo letras → casi seguro iniciales
    # Excluye 'pH' (mixed case, válido) y otros labels mixtos legítimos.
    if len(label) <= 3 and label.isupper() and label.isalpha():
        return False
    return True


def format_value(value) -> str:
    """Convierte un valor de celda a string limpio."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_pairs(ws) -> dict[str, str]:
    """
    Devuelve un dict {label: value} con las propiedades de la hoja.

    Estrategia: una fila puede tener label y valor con N columnas de gap entre
    medio (las plantillas industriales suelen tener merges/celdas vacías como
    separación visual). Por cada fila:
        1. Busca el último valor no vacío (de derecha a izquierda) = candidato VALUE.
        2. Escanea de derecha a izquierda desde ahí hasta col 0, buscando el primer
           string que pase is_valid_label() → ese es el LABEL.
    Si el mismo label aparece dos veces, conserva el primer valor encontrado.
    """
    pairs: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not cells:
            continue

        # 1) Encontrar el ÚLTIMO valor no vacío de la fila
        value_idx = None
        for i in range(len(cells) - 1, -1, -1):
            v = cells[i]
            if v is not None and format_value(v):
                value_idx = i
                break
        if value_idx is None or value_idx == 0:
            continue  # no hay value, o está en col 0 (no hay espacio para label)

        value_str = format_value(cells[value_idx])
        # Si el "value" es en realidad un header decorativo, descartar
        if value_str.upper() in IGNORE_LABELS:
            continue

        # 2) Escanear hacia la izquierda buscando un label válido
        for j in range(value_idx - 1, -1, -1):
            cand = cells[j]
            if cand is None or not isinstance(cand, str):
                continue
            label = normalize_label(cand)
            if not is_valid_label(label):
                continue
            if label not in pairs:
                pairs[label] = value_str
            break  # solo el label más cercano a la izquierda del value
    return pairs


# ─────────────────────────────────────────────────────────────────────────
# Extracción de imágenes embebidas en el Excel
# ─────────────────────────────────────────────────────────────────────────
_XLSX_NS = {
    'r':    'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'xdr':  'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a':    'http://schemas.openxmlformats.org/drawingml/2006/main',
}

# Filtro de logos por posición + tamaño:
#   Si el anchor superior izquierdo está en filas Excel 1-3 (0-based 0-2) Y
#   la imagen mide ≤ LOGO_MAX_HEIGHT_ROWS filas de alto → se considera logo
#   del header y se descarta.
#   Si está arriba pero mide más → es una foto real del equipo (alguien la
#   colocó en posición inusual) y se conserva.
HEADER_ANCHOR_MAX_ROW = 2       # filas 1-3 Excel
LOGO_MAX_HEIGHT_ROWS = 5        # un logo típico abarca 1-3 filas; una foto >10


def _resolve_target(base: str, target: str) -> str:
    """Resuelve una ruta relativa estilo '../media/image1.png' contra una base."""
    target = target.lstrip('/')
    if target.startswith('../'):
        # Subir niveles desde 'base' (que es la carpeta del archivo origen)
        base_parts = base.split('/')[:-1]  # carpeta
        while target.startswith('../'):
            target = target[3:]
            if base_parts:
                base_parts.pop()
        return '/'.join(base_parts + [target])
    return '/'.join(base.split('/')[:-1] + [target])


def _parse_drawing_anchors(zf: zipfile.ZipFile, drawing_path: str,
                            rid_to_target: dict[str, str]) -> list[tuple[str, int, int]]:
    """Parsea xl/drawings/drawing*.xml y devuelve [(image_path, from_row, to_row)].

    Ambas filas son 0-based.
      - from_row: fila de la esquina superior izquierda. -1 si absoluteAnchor.
      - to_row: fila de la esquina inferior derecha. -1 si oneCellAnchor o
        absoluteAnchor (no llevan `<xdr:to>`).
    """
    XDR = '{%s}' % _XLSX_NS['xdr']
    A = '{%s}' % _XLSX_NS['a']
    R = '{%s}' % _XLSX_NS['r']
    drawing_xml = ET.fromstring(zf.read(drawing_path))
    result: list[tuple[str, int, int]] = []

    def _row_of(elem):
        if elem is None:
            return -1
        row_e = elem.find(XDR + 'row')
        if row_e is None or row_e.text is None:
            return -1
        try:
            return int(row_e.text)
        except ValueError:
            return -1

    for tag in ('twoCellAnchor', 'oneCellAnchor', 'absoluteAnchor'):
        for anchor in drawing_xml.findall(XDR + tag):
            from_row = _row_of(anchor.find(XDR + 'from'))
            to_row = _row_of(anchor.find(XDR + 'to'))
            # Buscar TODOS los <a:blip> dentro del anchor a cualquier
            # profundidad — esto cubre:
            #   - <xdr:pic>/<xdr:blipFill>/<a:blip> (caso común)
            #   - <xdr:grpSp>/<xdr:pic>/<xdr:blipFill>/<a:blip> (group shapes
            #     con imagen, p. ej. fotos del equipo agrupadas con texto)
            #   - cualquier otro contenedor anidado
            for blip in anchor.iter(A + 'blip'):
                rid = blip.get(R + 'embed')
                if not rid or rid not in rid_to_target:
                    continue
                full_path = _resolve_target(drawing_path, rid_to_target[rid])
                result.append((full_path, from_row, to_row))
    return result


def _is_header_logo(from_row: int, to_row: int) -> bool:
    """True si la imagen parece logo del header (arriba + pequeña).

    Criterio: anclada en filas Excel 1-3 (0-based 0-2) Y mide pocas filas.
    Si el anchor no expone `to_row` (oneCellAnchor), asume logo cuando está arriba.
    """
    if not (0 <= from_row <= HEADER_ANCHOR_MAX_ROW):
        return False  # no está en el header
    if to_row < 0:
        return True   # oneCellAnchor en el header → asumir logo
    return (to_row - from_row) <= LOGO_MAX_HEIGHT_ROWS


def _sheet_to_images_map(zf: zipfile.ZipFile) -> dict[str, list[tuple[str, int, int]]]:
    """Devuelve {sheet_name: [(image_path, from_row, to_row), ...]} por hoja.

    Filas 0-based: 0=fila 1 Excel, 1=fila 2, etc.
    from_row=-1 si absoluteAnchor; to_row=-1 si oneCellAnchor/absoluteAnchor.
    """
    wb_xml = ET.fromstring(zf.read('xl/workbook.xml'))
    sheets = [{'name': s.get('name'),
               'rid': s.get('{%s}id' % _XLSX_NS['r'])}
              for s in wb_xml.find('main:sheets', _XLSX_NS)]
    wb_rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
    wb_rid_to_target = {r.get('Id'): r.get('Target') for r in wb_rels}

    result: dict[str, list[tuple[str, int]]] = {}
    for s in sheets:
        sheet_path = 'xl/' + wb_rid_to_target[s['rid']]
        rels_path = sheet_path.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        if rels_path not in zf.namelist():
            result[s['name']] = []
            continue
        sheet_rels = ET.fromstring(zf.read(rels_path))
        drawing_target = next(
            (r.get('Target') for r in sheet_rels
             if 'drawing' in r.get('Type', '').lower()),
            None,
        )
        if not drawing_target:
            result[s['name']] = []
            continue
        drawing_path = _resolve_target(sheet_path, drawing_target)
        drawing_rels_path = (
            drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
        )
        if drawing_rels_path not in zf.namelist():
            result[s['name']] = []
            continue
        drawing_rels = ET.fromstring(zf.read(drawing_rels_path))
        rid_to_target = {r.get('Id'): r.get('Target') for r in drawing_rels
                         if 'image' in r.get('Type', '').lower()}
        result[s['name']] = _parse_drawing_anchors(zf, drawing_path, rid_to_target)
    return result


def extract_images(input_path: Path, images_dir: Path) -> dict[str, list[str]]:
    """
    Extrae las imágenes específicas de cada hoja del Excel origen a
    images_dir/<TAG>/N.<ext> y devuelve un dict {TAG: [filenames]}.

    Dos filtros encadenados para descartar logos/plantillas:
      1. Filtro por POSICIÓN: imágenes con anchor en filas Excel 1-3
         (HEADER_ANCHOR_MAX_ROW = 2 en 0-based) se descartan — son logos
         del header del documento.
      2. Filtro por FRECUENCIA: imágenes presentes en > IMAGE_COMMON_THRESHOLD
         hojas se descartan — son plantillas decorativas comunes.

    Si una hoja queda sin imágenes después de los filtros, se aplica un
    fallback (conservar las del filtro #1) para que ningún TAG quede vacío.
    """
    print(f"🖼️  Extrayendo imágenes a: {images_dir}")
    images_dir.mkdir(parents=True, exist_ok=True)

    extracted: dict[str, list[str]] = {}
    with zipfile.ZipFile(input_path) as zf:
        # {sheet: [(image_path, anchor_row_0based), ...]}
        sheet_to_images = _sheet_to_images_map(zf)

        # FILTRO 1 — heurística posición + tamaño:
        #   - imagen anclada en filas 1-3 Y de pocas filas de alto → LOGO → descartar
        #   - imagen anclada arriba pero grande (>5 filas) → FOTO → conservar
        #   - imagen anclada bajo la fila 3 → no se evalúa, se conserva
        header_filtered: dict[str, list[str]] = {}
        header_logo_count = 0
        for sheet_name, anchored in sheet_to_images.items():
            kept = []
            for img_path, from_row, to_row in anchored:
                if _is_header_logo(from_row, to_row):
                    header_logo_count += 1
                    continue
                kept.append(img_path)
            header_filtered[sheet_name] = kept
        if header_logo_count:
            print(f"   ⊘ {header_logo_count} imágenes descartadas como logos del header "
                  f"(filas 1-{HEADER_ANCHOR_MAX_ROW + 1} + ≤{LOGO_MAX_HEIGHT_ROWS} filas alto)")

        # FILTRO 2 — por FRECUENCIA: descartar imágenes comunes a muchas hojas
        usage = defaultdict(int)
        for imgs in header_filtered.values():
            for img in imgs:
                usage[img] += 1

        for sheet_name, image_paths in header_filtered.items():
            specific = [p for p in image_paths
                       if usage[p] <= IMAGE_COMMON_THRESHOLD]
            # Fallback: si TODAS las imágenes de la hoja superan el umbral
            # de frecuencia, conservar las que pasaron el filtro #1
            # (mejor mostrar imagen reutilizada que ninguna)
            if not specific and image_paths:
                print(f"   ℹ️  '{sheet_name}': todas las imágenes son 'comunes'; "
                      f"se conservan {len(image_paths)} igual (fallback)")
                specific = list(image_paths)
            if not specific:
                continue
            # Expandir el nombre de hoja a TAGs (igual que en process_workbook)
            tags = expand_tags_from_sheet_name(sheet_name)
            for tag in tags:
                # Slug del TAG (mismo que en excel_migrator.Item.filename)
                slug = tag.replace(' ', '').replace('/', '_')
                tag_dir = images_dir / slug
                tag_dir.mkdir(parents=True, exist_ok=True)
                filenames: list[str] = []
                for i, img_path in enumerate(specific, start=1):
                    if img_path not in zf.namelist():
                        continue
                    ext = Path(img_path).suffix.lower()  # .png / .jpeg
                    out_name = f"{i}{ext}"
                    out_file = tag_dir / out_name
                    with zf.open(img_path) as src, open(out_file, 'wb') as dst:
                        shutil.copyfileobj(src, dst)
                    filenames.append(out_name)
                if filenames:
                    extracted[slug] = filenames

    # Escribir manifest para que el migrador lo lea
    manifest = {tag: files for tag, files in sorted(extracted.items())}
    manifest_path = images_dir / "manifest.json"
    import json
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"   ✅ {len(extracted)} TAGs con imágenes, manifest.json escrito")
    return extracted


# ─────────────────────────────────────────────────────────────────────────
# Procesamiento del Excel origen
# ─────────────────────────────────────────────────────────────────────────
def process_workbook(input_path: Path) -> tuple[list[dict], list[str]]:
    """
    Lee el Excel de fichas y devuelve:
        - rows: lista de dicts {TAG, Categoría, Título, <propiedades>}
        - prop_columns: lista ordenada de columnas de propiedades detectadas
    """
    print(f"📖 Leyendo: {input_path}")
    wb = load_workbook(input_path, data_only=True)

    rows: list[dict] = []
    all_props: list[str] = []  # mantiene el orden de aparición
    seen_props: set[str] = set()

    # Patrón que valida nombre de hoja como TAG (rechaza "Hoja3", "Sheet1", etc.)
    valid_tag_re = re.compile(r"^\d{3}-[A-Z]+-")

    for sheet_name in wb.sheetnames:
        # Filtrar hojas-ruido: Excel pone "Hoja3", "Sheet1" como default
        if not valid_tag_re.match(sheet_name.strip()):
            print(f"  ⊘ {sheet_name:30} → ignorada (no parece TAG: 'NNN-TYPE-...')")
            continue

        ws = wb[sheet_name]
        tags = expand_tags_from_sheet_name(sheet_name)
        pairs = extract_pairs(ws)

        # Filtrar hojas sin propiedades reales (ruido)
        if not pairs:
            print(f"  ⊘ {sheet_name:30} → ignorada (0 propiedades extraídas)")
            continue

        category = infer_category(tags[0])

        # Acumular orden de aparición de las propiedades (preserva orden del Excel origen)
        for k in pairs:
            if k not in seen_props:
                seen_props.add(k)
                all_props.append(k)

        for tag in tags:
            row = {"TAG": tag, "Categoría": category, "Título": ""}
            row.update(pairs)
            rows.append(row)

        print(f"  ✓ {sheet_name:30} → {len(tags)} TAG(s) [{category:13}] · {len(pairs)} propiedades")

    return rows, all_props


# ─────────────────────────────────────────────────────────────────────────
# Generación del Excel destino
# ─────────────────────────────────────────────────────────────────────────
def style_header(ws, n_cols):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0969DA")
    border = Border(bottom=Side(style="medium", color="0550AE"))
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 36


def autosize_columns(ws, headers, rows):
    for i, h in enumerate(headers, start=1):
        max_len = max(len(h), 12)
        for r in rows:
            v = str(r.get(h, "") or "")
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 3, 14), 50)


def add_category_validation(ws, n_rows):
    formula = f'"{",".join(CATEGORIES)}"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=False,
        showDropDown=False,
        errorTitle="Categoría inválida",
        error=f"Debe ser una de: {', '.join(CATEGORIES)}",
        promptTitle="Categoría",
        prompt=f"Selecciona: {', '.join(CATEGORIES)}",
    )
    ws.add_data_validation(dv)
    last_row = max(n_rows + 100, 500)
    dv.add(f"B2:B{last_row}")


def write_data_sheet(wb: Workbook, rows: list[dict], prop_columns: list[str]):
    ws = wb.active
    ws.title = "Datos"

    headers = ["TAG", "Categoría", "Título"] + prop_columns + ["Ficha Técnica", "Curva"]

    # Headers
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    style_header(ws, len(headers))

    # Filas
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            v = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=v if v else None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "C2"
    add_category_validation(ws, len(rows))
    autosize_columns(ws, headers, rows)


def write_instructions_sheet(wb: Workbook):
    ws = wb.create_sheet("Instrucciones")
    title_font = Font(bold=True, size=14, color="0969DA")
    h2_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    code_font = Font(size=10, name="Consolas", color="333333")

    lines = [
        ("Plantilla de datos para el sitio QR Groupe SEB", title_font),
        ("", normal_font),
        ("Cada fila es un TAG. Cada columna a partir de 'Título' es una propiedad", normal_font),
        ("que aparecerá como sección colapsable en la ficha del TAG.", normal_font),
        ("Si dejas una celda vacía, esa sección NO se genera en la ficha.", normal_font),
        ("", normal_font),
        ("Columnas obligatorias", h2_font),
        ("• TAG          identificador único (ej. 100-P-01A)", normal_font),
        ("• Categoría    EQUIPOS, INSTRUMENTOS o TANQUES (dropdown en columna B)", normal_font),
        ("", normal_font),
        ("Columnas opcionales", h2_font),
        ("• Título           si lo dejas vacío, se usa el TAG como título", normal_font),
        ("• <propiedades>    cada columna detectada del Excel origen", normal_font),
        ("• Ficha Técnica    inserta hyperlink con Ctrl+K → pega URL de Drive", normal_font),
        ("• Curva            ídem", normal_font),
        ("", normal_font),
        ("Cómo regenerar este Excel desde fichas técnicas", h2_font),
        ("Si recibes un Excel con una hoja por TAG (formato ficha técnica), ejecuta:", normal_font),
        ("    python3 convert_fichas_to_template.py <archivo.xlsx>", code_font),
        ("Se sobrescribirá esta plantilla con los datos del nuevo archivo.", normal_font),
        ("", normal_font),
        ("Cómo regenerar el sitio desde este Excel", h2_font),
        ("    ./update_site_from_excel.sh         # regenera docs/ sin pushear", code_font),
        ("    ./update_site_from_excel.sh --push  # regenera y publica a GitHub Pages", code_font),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = font
        c.alignment = Alignment(vertical="top")
    ws.column_dimensions["A"].width = 95
    ws.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}
    input_path = Path(args[0]) if args else INPUT_DEFAULT
    if not input_path.exists():
        print(f"❌ No existe: {input_path}", file=sys.stderr)
        sys.exit(1)

    # --images-only: solo re-extrae imágenes, no regenera plantilla_sitio.xlsx
    # (útil si plantilla está editada a mano y no querés perder los hyperlinks)
    if "--images-only" in flags:
        print(f"🖼️  Modo --images-only: solo extrae imágenes desde {input_path.name}")
        extract_images(input_path, IMAGES_DIR)
        return

    rows, prop_columns = process_workbook(input_path)

    # Extraer imágenes embebidas del Excel origen a docs/assets/images/
    print()
    extract_images(input_path, IMAGES_DIR)
    print()

    wb = Workbook()
    write_data_sheet(wb, rows, prop_columns)
    write_instructions_sheet(wb)
    wb.save(OUTPUT)

    print()
    print(f"✅ Plantilla generada: {OUTPUT}")
    print(f"   • {len(rows)} TAGs en total")
    print(f"   • {len(prop_columns)} columnas de propiedades + Ficha Técnica + Curva")
    print()
    print("Categorías:")
    from collections import Counter
    counts = Counter(r["Categoría"] for r in rows)
    for cat in CATEGORIES:
        print(f"  {cat:13} {counts.get(cat, 0)}")


if __name__ == "__main__":
    main()
