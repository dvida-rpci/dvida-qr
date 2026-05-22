#!/usr/bin/env python3
"""
generate_tag_resources_template.py
==================================
Genera (o actualiza) TAG_RESOURCES.xlsx con una hoja `TAG_RESOURCES` y
columnas: TAG_ID, Specifications, Handbook_Manual, Maintenance.

- Si el archivo NO existe → lo crea con todos los TAGs de plantilla_sitio.xlsx
  y celdas de recurso vacías (las llena el usuario con Ctrl+K).
- Si el archivo YA existe → preserva los hyperlinks existentes y solo agrega
  filas para TAGs nuevos que aún no estuvieran.

Uso:
    python3 generate_tag_resources_template.py
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).parent
PLANTILLA = REPO_ROOT / "plantilla_sitio.xlsx"
OUTPUT = REPO_ROOT / "TAG_RESOURCES.xlsx"
SHEET = "TAG_RESOURCES"
HEADERS = ["TAG_ID", "Specifications", "Handbook_Manual", "Maintenance"]


def read_tags_from_plantilla() -> list[str]:
    if not PLANTILLA.exists():
        raise FileNotFoundError(f"No existe {PLANTILLA}")
    wb = load_workbook(PLANTILLA, data_only=True)
    ws = wb["Datos"] if "Datos" in wb.sheetnames else wb.active
    tags = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tag = row[0]
        if tag:
            tags.append(str(tag).strip())
    return tags


def style_headers(ws):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0969DA")
    border = Border(bottom=Side(style="medium", color="0550AE"))
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 20
    for col_letter in ("B", "C", "D"):
        ws.column_dimensions[col_letter].width = 30
    ws.freeze_panes = "B2"


def create_new_workbook(tags: list[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for col_idx, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    style_headers(ws)
    for tag in tags:
        ws.append([tag, "", "", ""])
    return wb


def merge_into_existing(tags: list[str]):
    wb = load_workbook(OUTPUT)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    existing_tags = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing_tags.add(str(row[0]).strip())
    missing = [t for t in tags if t not in existing_tags]
    for tag in missing:
        ws.append([tag, "", "", ""])
    return wb, len(missing)


def main():
    tags = read_tags_from_plantilla()
    print(f"📖 {len(tags)} TAGs leídos de plantilla_sitio.xlsx")

    if OUTPUT.exists():
        wb, added = merge_into_existing(tags)
        if added:
            print(f"   ➕ {added} TAGs nuevos agregados a TAG_RESOURCES.xlsx (existentes preservados)")
        else:
            print(f"   ✔ TAG_RESOURCES.xlsx ya tiene todos los TAGs; no se modifica")
    else:
        wb = create_new_workbook(tags)
        print(f"   🆕 Creando TAG_RESOURCES.xlsx desde cero")

    wb.save(OUTPUT)
    print(f"✅ {OUTPUT}")
    print("   Llena las celdas de Specifications / Handbook_Manual / Maintenance")
    print("   con hyperlinks (Ctrl+K) y vuelve a correr ./update_site_from_excel.sh")


if __name__ == "__main__":
    main()
